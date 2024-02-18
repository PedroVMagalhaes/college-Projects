import openpyxl
import os

def money(number, n_people):
    # Get the path of the current script
    script_path = os.path.dirname(os.path.abspath(__file__))

    # Navigate to the parent directory
    parent_directory = os.path.join(script_path, '..')

    # Construct the file paths
    weekly_file_path = os.path.join(parent_directory, 'weekly.xlsx')
    names_file_path = os.path.join(parent_directory, 'roster.xlsx')

    # Load the existing excel's
    workbook = openpyxl.load_workbook(weekly_file_path)
    names_workbook = openpyxl.load_workbook(names_file_path)

    # Sheets assignemt
    worksheet = workbook.active
    names_worksheet = names_workbook.active

    # Read the existing value in A1 and B1
    existing_value = worksheet['A1'].value
    old_week = worksheet['B1'].value

    # Increment the value and week
    new_value = existing_value + 3
    week = old_week + 1

    # Input
    input_number = float(number)
    number_of_people = n_people
    breaker = 0
    
    # Cuts amount 80/20
    amount_for_people = (0.8 * input_number) / number_of_people
    amount_for_management = 0.2 * input_number

    # Update the value in A1 and B1.
    worksheet['A1'] = new_value
    worksheet['B1'] = week
    ind = 4
    #First Iteration
    if old_week == 0:
        worksheet.cell(row=2, column=1, value="Week")
        worksheet.cell(row=2, column=2, value=week)

        # Assign names from names worksheet (columns A to E, rows 1 to 5)
        for col_num in range(1, 6):
            for row_num in range(1, 6):
                name_cell = names_worksheet.cell(row=row_num, column=col_num).value
                worksheet.cell(row=ind, column=1, value=name_cell)
                worksheet.cell(row=ind, column=2, value=amount_for_people)
                ind = ind +1
                breaker = breaker + 1
                if number_of_people == breaker:
                    break 
            if number_of_people == breaker:
                    break          
        
        #Add Management part at the end
        worksheet.cell(row=31, column=1 , value="management") 
        worksheet.cell(row=31, column=2 , value=amount_for_management)     

        # Save the changes
        workbook.save(weekly_file_path)
    
    #All other Iterations
    else:
        worksheet.cell(row=2, column=existing_value , value="Week")
        worksheet.cell(row=2, column=existing_value + 1, value=week)

        # Assign names from names worksheet (columns A to E, rows 1 to 5)
        for col_num in range(1, 6):
            for row_num in range(1, 6):
                name_cell = names_worksheet.cell(row=row_num, column=col_num).value
                worksheet.cell(row=ind, column=existing_value, value=name_cell)
                worksheet.cell(row=ind, column=existing_value + 1, value=amount_for_people)
                ind = ind +1
                breaker = breaker + 1
                if number_of_people == breaker:
                    break 
            if number_of_people == breaker:
                    break   

        #Add Management part at the end
        worksheet.cell(row=31, column=existing_value , value="management") 
        worksheet.cell(row=31, column=existing_value + 1 , value=amount_for_management)      

        # Save the changes
        workbook.save(weekly_file_path)

