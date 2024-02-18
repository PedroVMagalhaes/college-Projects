import openpyxl
import sys

def main():
    file_path = "weekly.xlsx"

    # Load the existing workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the first sheet (assuming it's the default sheet)
    worksheet = workbook.active

    # Read the existing value in A1
    existing_value = worksheet['A1'].value
    old_week = worksheet['B1'].value

    # Increment the value
    new_value = existing_value + 3

    # Increment the value
    week = old_week + 1

    # Prompt the user for input
    input_number = float(input("Enter a number: "))
    
    # Calculate amounts for people and management
    amount_for_people = (0.8 * input_number) / 25
    amount_for_management = 0.2 * input_number

    # Update the value in A1
    worksheet['A1'] = new_value
    worksheet['B1'] = week

    if old_week == 0:
        worksheet.cell(row=2, column=1, value="Week")
        worksheet.cell(row=2, column=2, value=week)

        # Assign names to people
        people_names = [f"Person {i}" for i in range(1, 26)]

        # Distribute the amount among 25 people
        for row_num, person_name in enumerate(people_names, start=4):
            worksheet.cell(row=row_num, column=1, value=person_name)
            worksheet.cell(row=row_num, column=2, value=amount_for_people)
        
        worksheet.cell(row=29, column=1 , value="management") 
        worksheet.cell(row=29, column=2 , value=amount_for_management)     

        # Save the changes
        workbook.save(file_path)

    else:
        worksheet.cell(row=2, column=existing_value , value="Week")
        worksheet.cell(row=2, column=existing_value + 1, value=week)


        # Assign names to people
        people_names = [f"Person {i}" for i in range(1, 26)]

        # Distribute the amount among 25 people
        for row_num, person_name in enumerate(people_names, start=4):
            worksheet.cell(row=row_num, column=existing_value, value=person_name)
            worksheet.cell(row=row_num, column=existing_value +1, value=amount_for_people)
    
        worksheet.cell(row=29, column=existing_value , value="management") 
        worksheet.cell(row=29, column=existing_value +1 , value=amount_for_management)      

        # Save the changes
        workbook.save(file_path)

if __name__ == "__main__":
    main()
